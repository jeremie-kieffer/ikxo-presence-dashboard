"""
Migration des données du xlsx vers Supabase.

Lit public/data/suivi_presence_consultants.xlsx et insère dans les 6 tables Supabase
en respectant l'ordre des foreign keys.

Idempotent : peut être re-exécuté sans créer de doublons grâce aux upserts
sur les contraintes UNIQUE/PK de chaque table.

Usage:
    1. Avoir un fichier .env à la racine du repo contenant :
        SUPABASE_URL=https://xxxxxxxx.supabase.co
        SUPABASE_SERVICE_ROLE_KEY=eyJ...
    2. pip install supabase python-dotenv --user
    3. Depuis la racine du repo : python3 scripts/migrate_xlsx_to_supabase.py
"""

import os
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook
from dotenv import load_dotenv
from supabase import create_client


# ============================================================
# Setup
# ============================================================
load_dotenv()
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    raise SystemExit(
        "⛔ Variables SUPABASE_URL et SUPABASE_SERVICE_ROLE_KEY manquantes. "
        "Vérifie ton fichier .env à la racine du repo."
    )

REPO_ROOT = Path(__file__).parent.parent
XLSX_PATH = REPO_ROOT / 'public' / 'data' / 'suivi_presence_consultants.xlsx'

if not XLSX_PATH.exists():
    raise SystemExit(f"⛔ Fichier xlsx introuvable : {XLSX_PATH}")

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

# Test connexion (fail fast si auth KO)
try:
    sb.table('consultants').select('id').limit(1).execute()
except Exception as e:
    raise SystemExit(f"⛔ Connexion Supabase KO : {e}")

print(f"🔗 Connecté à {SUPABASE_URL}")
print(f"📂 Lecture du fichier : {XLSX_PATH.name}\n")

wb = load_workbook(XLSX_PATH, data_only=True)
CHUNK = 500  # taille de batch pour les inserts/upserts


# ============================================================
# Helpers de conversion
# ============================================================
def to_date(value):
    """Convertit une valeur Excel en date Python (ou None)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Cas serial Excel : nombre de jours depuis 1900-01-01
        # (les cellules sans format date sont lues comme int/float)
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def to_str_date(d):
    """Sérialise une date Python en chaîne ISO (ou None)."""
    return d.isoformat() if d else None


def trouver_onglet(nom_cible):
    """Trouve un onglet en tolérant les accents et espaces."""
    cible_norm = nom_cible.lower().replace('é', 'e').replace('è', 'e').strip()
    for name in wb.sheetnames:
        norm = name.lower().replace('é', 'e').replace('è', 'e').strip()
        if norm == cible_norm:
            return wb[name]
    raise ValueError(f"Onglet '{nom_cible}' introuvable")


# ============================================================
# 1. CONSULTANTS
# ============================================================
print("=== 1. Migration consultants ===")
ref = trouver_onglet('Référentiel')

# Existants en DB : map nom → uuid (pour idempotence)
existing = {row['nom']: row['id'] for row in sb.table('consultants').select('id, nom').execute().data}

a_inserer = []
for r in range(4, 60):
    nom = ref.cell(r, 1).value
    if not nom or not isinstance(nom, str) or nom.strip() == '':
        continue
    nom = nom.strip()
    if nom in existing:
        continue
    role_brut = ref.cell(r, 4).value
    role = (role_brut or 'consultant').lower().strip()
    if role not in ('consultant', 'interne'):
        role = 'consultant'
    a_inserer.append({
        'nom': nom,
        'date_entree': to_str_date(to_date(ref.cell(r, 2).value)),
        'date_sortie': to_str_date(to_date(ref.cell(r, 3).value)),
        'role': role,
    })

if a_inserer:
    result = sb.table('consultants').insert(a_inserer).execute()
    for c in result.data:
        existing[c['nom']] = c['id']
    print(f"  ✓ {len(result.data)} consultants insérés")
else:
    print("  – Aucun nouveau consultant")

nom_to_id = existing  # map nom → uuid pour la suite
print(f"  → {len(nom_to_id)} consultants au total en DB\n")


# ============================================================
# 2. SESSIONS_FORMATION
# ============================================================
print("=== 2. Migration sessions_formation ===")
formations = trouver_onglet('Formations')

a_upserter = []
for r in range(4, 50):
    sid = formations.cell(r, 1).value
    if not sid or not isinstance(sid, str):
        continue
    a_upserter.append({
        'id': sid.strip(),
        'date': to_str_date(to_date(formations.cell(r, 2).value)),
        'thematique': (formations.cell(r, 3).value or '').strip(),
        'lien_support': formations.cell(r, 5).value,
    })

if a_upserter:
    result = sb.table('sessions_formation').upsert(a_upserter, on_conflict='id').execute()
    print(f"  ✓ {len(result.data)} sessions upsertées\n")
else:
    print("  – Aucune session trouvée\n")


# ============================================================
# 3. EVENEMENTS
# ============================================================
print("=== 3. Migration evenements ===")
try:
    ev = trouver_onglet('Événements')
except ValueError:
    ev = None

if ev:
    existing_events = set()
    for row in sb.table('evenements').select('date, libelle').execute().data:
        existing_events.add((row['date'], row['libelle']))

    a_inserer = []
    # Structure de l'onglet Événements :
    #   ligne 1 = titre fusionné
    #   ligne 2 = en-têtes
    #   ligne 3+ = données
    #   col A = Date, col B = Type, col C = Libellé / Détail
    for r in range(3, 50):
        d = to_date(ev.cell(r, 1).value)
        type_val = ev.cell(r, 2).value
        libelle = ev.cell(r, 3).value
        if not d or not libelle:
            continue
        if (d.isoformat(), libelle) in existing_events:
            continue
        a_inserer.append({
            'date': d.isoformat(),
            'libelle': libelle,
            'type': type_val,
        })

    if a_inserer:
        result = sb.table('evenements').insert(a_inserer).execute()
        print(f"  ✓ {len(result.data)} événements insérés\n")
    else:
        print("  – Aucun nouvel événement\n")
else:
    print("  ⚠ Onglet Événements introuvable, skipped\n")


# ============================================================
# 4. PRESENCES (matrices Saisie YYYY-MM → format long)
# ============================================================
print("=== 4. Migration presences (matrices → long format) ===")
CODES_PRESENCE = {
    '1': 'present',
    'IC': 'intercontract',
    'M': 'absence_longue',
}

presences = []
for sheet_name in wb.sheetnames:
    if not sheet_name.lower().startswith('saisie'):
        continue
    s = wb[sheet_name]

    # Ligne 3 = en-têtes de dates en colonnes C+
    dates_col = {}
    for c in range(3, s.max_column + 1):
        d = to_date(s.cell(3, c).value)
        if d:
            dates_col[c] = d

    # Lignes 4+ : col A = nom du consultant, col C+ = codes
    for r in range(4, s.max_row + 1):
        nom = s.cell(r, 1).value
        if not nom or not isinstance(nom, str):
            continue
        nom = nom.strip()
        if nom not in nom_to_id:
            continue
        consultant_id = nom_to_id[nom]
        for c, d in dates_col.items():
            code = s.cell(r, c).value
            if not code:
                continue
            code_str = str(code).strip().upper()
            if code_str == 'X':
                # Jour d'événement : décision produit = pas de ligne presence
                continue
            statut = CODES_PRESENCE.get(code_str)
            if not statut:
                continue
            presences.append({
                'consultant_id': consultant_id,
                'date': d.isoformat(),
                'statut': statut,
            })

n_upserted = 0
for i in range(0, len(presences), CHUNK):
    chunk = presences[i:i + CHUNK]
    result = sb.table('presences').upsert(chunk, on_conflict='consultant_id,date').execute()
    n_upserted += len(result.data)
print(f"  ✓ {n_upserted} présences upsertées (sur {len(presences)} candidates)\n")


# ============================================================
# 5. PARTICIPATIONS_FORMATION (matrice → format long)
# ============================================================
print("=== 5. Migration participations_formation ===")
CODES_PARTICIPATION = {
    'P': 'present',
    'F': 'formateur',
    'I': 'inscrit',
}

fp = trouver_onglet('Formations_Participations')

# Ligne 3 = ids de session à partir de col C
sessions_col = {}
for c in range(3, fp.max_column + 1):
    sid = fp.cell(3, c).value
    if sid and isinstance(sid, str) and sid.startswith('F-'):
        sessions_col[c] = sid.strip()

participations = []
for r in range(4, fp.max_row + 1):
    nom = fp.cell(r, 1).value
    if not nom or not isinstance(nom, str):
        continue
    nom = nom.strip()
    if nom not in nom_to_id:
        continue
    consultant_id = nom_to_id[nom]
    for c, sid in sessions_col.items():
        code = fp.cell(r, c).value
        if not code:
            continue
        statut = CODES_PARTICIPATION.get(str(code).strip().upper())
        if not statut:
            continue
        participations.append({
            'session_id': sid,
            'consultant_id': consultant_id,
            'statut': statut,
        })

n_upserted = 0
for i in range(0, len(participations), CHUNK):
    chunk = participations[i:i + CHUNK]
    result = sb.table('participations_formation').upsert(
        chunk, on_conflict='session_id,consultant_id'
    ).execute()
    n_upserted += len(result.data)
print(f"  ✓ {n_upserted} participations upsertées (sur {len(participations)} candidates)\n")


# ============================================================
# 6. FEEDBACKS_FORMATION
# ============================================================
print("=== 6. Migration feedbacks_formation ===")
fb = trouver_onglet('Formation_Feedbacks')

feedbacks = []
for r in range(4, fb.max_row + 1):
    rid = fb.cell(r, 1).value
    sid = fb.cell(r, 2).value
    ts = fb.cell(r, 3).value
    note = fb.cell(r, 4).value
    application = fb.cell(r, 5).value
    if not rid or not sid or not isinstance(ts, datetime) or note is None:
        continue
    feedbacks.append({
        'id': rid.strip(),
        'session_id': sid.strip(),
        'timestamp': ts.isoformat(),
        'note_globale': int(note),
        'application': (application or '').strip(),
        'verbatim_apprecie': fb.cell(r, 6).value,
        'verbatim_amelioration': fb.cell(r, 7).value,
        'verbatim_commentaire': fb.cell(r, 8).value,
    })

n_upserted = 0
for i in range(0, len(feedbacks), CHUNK):
    chunk = feedbacks[i:i + CHUNK]
    result = sb.table('feedbacks_formation').upsert(chunk, on_conflict='id').execute()
    n_upserted += len(result.data)
print(f"  ✓ {n_upserted} feedbacks upsertés (sur {len(feedbacks)} candidats)\n")


print("✅ Migration terminée.")
