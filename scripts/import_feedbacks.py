#!/usr/bin/env python3
"""
Import des feedbacks Formation depuis CSV Google Forms vers Excel.

Usage : python scripts/import_feedbacks.py

Lit tous les fichiers feedback_F-YYYY-NNN.csv du dossier imports_feedbacks/
et ajoute leurs lignes à l'onglet Formation_Feedbacks du fichier xlsx,
chaque ligne taggée avec son id_session extrait du nom de fichier.

Idempotent : ne ré-importe pas une ligne déjà présente (matching sur
id_session + timestamp). Tu peux relancer le script autant de fois que
tu veux sans créer de doublons.

Convention de nommage des CSV : feedback_F-YYYY-NNN.csv
  exemple : feedback_F-2025-001.csv
"""
import csv
import re
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

# --- Chemins (calculés depuis l'emplacement du script) ---
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
XLSX_PATH = REPO_ROOT / "public" / "data" / "suivi_presence_consultants.xlsx"
IMPORTS_DIR = REPO_ROOT / "imports_feedbacks"
SHEET_NAME = "Formation_Feedbacks"

# --- Patterns pour matcher les colonnes du CSV Google Forms ---
# Le script utilise des regex pour identifier chaque colonne dans le CSV,
# ce qui le rend tolérant aux petites variations entre tes 8 forms historiques.
COLUMN_PATTERNS = {
    "timestamp":             r"horodat|timestamp",
    "note_globale":          r"échelle de 1 à 5|évalues",
    "application":           r"appliquer rapidement|appris",
    "verbatim_apprecie":     r"apprécié|aimé",
    "verbatim_amelioration": r"amélior",
    "verbatim_commentaire":  r"autre commentaire|à nous partager",
}


def find_column(headers, pattern):
    """Retourne l'index de la première colonne dont l'en-tête matche le pattern."""
    rx = re.compile(pattern, re.IGNORECASE)
    for i, h in enumerate(headers):
        if h and rx.search(h):
            return i
    return None


def extract_session_id(filename):
    """Extrait F-YYYY-NNN depuis un nom de fichier comme 'feedback_F-2025-001.csv'."""
    m = re.search(r"(F-\d{4}-\d{3})", filename)
    return m.group(1) if m else None


def parse_timestamp(s):
    """Parse un timestamp Google Forms (plusieurs formats possibles selon la locale)."""
    if not s or not s.strip():
        return None
    s = s.strip()
    # Google Forms peut suffixer la timezone (« UTC+3 », « GMT-5 »…).
    # strptime ne sait pas la parser proprement, et on n'en a pas besoin
    # ici : on garde le timestamp tel qu'affiché côté form.
    s = re.sub(r"\s+(UTC|GMT)[+-]?\d{1,2}(:\d{2})?$", "", s)
    for fmt in (
        "%d/%m/%Y %H:%M:%S",  # FR
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",  # US
        "%d/%m/%Y %H:%M",
        "%Y/%m/%d %I:%M:%S %p",  # Google Forms 12h (vu sur F-2026-009)
        "%Y/%m/%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def find_next_empty_row(ws, start=4):
    """Première ligne vide à partir de start (1-indexé)."""
    r = start
    while ws.cell(r, 1).value is not None:
        r += 1
    return r


def get_max_response_num(ws):
    """Retourne le plus grand numéro de r-XXX déjà dans l'onglet."""
    n = 0
    for r in range(4, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if rid and isinstance(rid, str) and rid.startswith("r-"):
            try:
                n = max(n, int(rid[2:]))
            except ValueError:
                pass
    return n


def main():
    if not XLSX_PATH.exists():
        print(f"❌ Fichier introuvable : {XLSX_PATH}")
        sys.exit(1)

    if not IMPORTS_DIR.exists():
        print(f"⚠ Dossier {IMPORTS_DIR} n'existe pas — création.")
        IMPORTS_DIR.mkdir(parents=True, exist_ok=True)

    csv_files = sorted(IMPORTS_DIR.glob("feedback_F-*.csv"))
    if not csv_files:
        print(f"Aucun feedback_F-*.csv dans {IMPORTS_DIR}. Rien à importer.")
        sys.exit(0)

    print(f"📥 {len(csv_files)} fichier(s) CSV trouvé(s).\n")

    wb = load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Onglet {SHEET_NAME} introuvable dans le xlsx.")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    # Set des feedbacks déjà importés (clé = id_session + timestamp en string)
    existing = set()
    for r in range(4, ws.max_row + 1):
        sid = ws.cell(r, 2).value
        ts = ws.cell(r, 3).value
        if sid and ts:
            existing.add((sid, str(ts)))

    next_num = get_max_response_num(ws) + 1
    total_imported = 0
    total_skipped = 0

    for csv_file in csv_files:
        sid = extract_session_id(csv_file.name)
        if not sid:
            print(f"  ⚠ Skip {csv_file.name} : pas de F-YYYY-NNN dans le nom.")
            continue

        with open(csv_file, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if not headers:
                print(f"  ⚠ Skip {csv_file.name} : fichier vide.")
                continue

            col_map = {k: find_column(headers, p) for k, p in COLUMN_PATTERNS.items()}

            missing = [k for k in ("timestamp", "note_globale") if col_map[k] is None]
            if missing:
                print(f"  ⚠ Skip {csv_file.name} : colonnes manquantes {missing}")
                print(f"     En-têtes vues : {headers}")
                continue

            imported_here = 0
            skipped_here = 0

            for row in reader:
                if not any(row):
                    continue

                ts_raw = row[col_map["timestamp"]] if col_map["timestamp"] < len(row) else ""
                ts = parse_timestamp(ts_raw)
                if not ts:
                    print(f"  ⚠ Ligne ignorée dans {csv_file.name} : timestamp invalide ({ts_raw!r})")
                    continue

                if (sid, str(ts)) in existing:
                    skipped_here += 1
                    continue

                note_raw = row[col_map["note_globale"]] if col_map["note_globale"] < len(row) else ""
                try:
                    note = int(note_raw)
                except (ValueError, TypeError):
                    note = None

                def cell_val(key):
                    idx = col_map[key]
                    if idx is None or idx >= len(row):
                        return ""
                    return row[idx].strip()

                target_row = find_next_empty_row(ws)
                rid = f"r-{next_num:03d}"
                next_num += 1

                ws.cell(target_row, 1, rid)
                ws.cell(target_row, 2, sid)
                c_ts = ws.cell(target_row, 3, ts)
                c_ts.number_format = "dd/mm/yyyy hh:mm"
                ws.cell(target_row, 4, note)
                ws.cell(target_row, 5, cell_val("application"))
                ws.cell(target_row, 6, cell_val("verbatim_apprecie"))
                ws.cell(target_row, 7, cell_val("verbatim_amelioration"))
                ws.cell(target_row, 8, cell_val("verbatim_commentaire"))

                existing.add((sid, str(ts)))
                imported_here += 1

            status = f"+{imported_here}" if imported_here else "0"
            if skipped_here:
                status += f" (dont {skipped_here} déjà présent{'s' if skipped_here > 1 else ''})"
            print(f"  ✓ {csv_file.name} [{sid}] → {status}")
            total_imported += imported_here
            total_skipped += skipped_here

    wb.save(XLSX_PATH)
    print(f"\n✅ Total : {total_imported} feedback(s) importé(s), {total_skipped} doublon(s) ignoré(s).")
    print(f"   Fichier sauvegardé : {XLSX_PATH.relative_to(REPO_ROOT)}")
    print(f"\nProchaine étape :")
    print(f"  git add public/data/suivi_presence_consultants.xlsx")
    print(f"  git commit -m \"Import feedbacks formation\"")
    print(f"  git push")


if __name__ == "__main__":
    main()
